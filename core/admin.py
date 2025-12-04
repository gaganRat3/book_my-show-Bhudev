from django.utils.html import format_html
from django.contrib import admin
# PaymentScreenshot admin customization
from .models import PaymentScreenshot

class PaymentScreenshotAdmin(admin.ModelAdmin):
	list_display = ('user_name', 'screenshot_thumbnail', 'uploaded_at')

	def user_name(self, obj):
		return obj.user.name if obj.user else "-"
	user_name.short_description = 'User Name'

	def screenshot_thumbnail(self, obj):
		if obj.image:
			return format_html('<img src="{}" width="100" style="object-fit:contain;" />', obj.image.url)
		return "No Image"
	screenshot_thumbnail.short_description = 'Screenshot'

admin.site.register(PaymentScreenshot, PaymentScreenshotAdmin)
from django.urls import path
from django.http import HttpResponse
from django.template import loader
from django.shortcuts import redirect
from core.models import SelectedSeat, LandingFormData

class SelectedSeatSummaryAdmin(admin.ModelAdmin):
	def get_urls(self):
		urls = super().get_urls()
		custom_urls = [
			path('selectedseat-summary/', self.admin_site.admin_view(self.selectedseat_summary_view), name="selectedseat-summary"),
			path('selectedseat-summary/edit/<int:user_id>/', self.admin_site.admin_view(self.edit_selectedseat_view), name="edit-selectedseat"),
			path('selectedseat-summary/export/', self.admin_site.admin_view(self.export_selectedseat_excel), name="export-selectedseat-excel"),
		]
		return custom_urls + urls

	def export_selectedseat_excel(self, request):
		import openpyxl
		from openpyxl.utils import get_column_letter
		from django.http import HttpResponse
		users = LandingFormData.objects.all()
		wb = openpyxl.Workbook()
		ws = wb.active
		ws.title = "Selected Seats"
		headers = ["User Name", "Phone", "Seats", "Total Paid", "Earliest Booking"]
		ws.append(headers)
		for user in users:
			seats = SelectedSeat.objects.filter(user=user)
			seat_labels = [s.seat.seat_number for s in seats]
			total_paid = sum(float(s.price or 0) for s in seats)
			earliest = seats.order_by('selected_at').first().selected_at if seats.exists() else None
			ws.append([
				user.name,
				user.phone,
				', '.join(seat_labels),
				total_paid,
				str(earliest) if earliest else ""
			])
		for col in range(1, len(headers) + 1):
			ws.column_dimensions[get_column_letter(col)].width = 20
		response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
		response['Content-Disposition'] = 'attachment; filename=selected_seat_summary.xlsx'
		wb.save(response)
		return response
	def edit_selectedseat_view(self, request, user_id):
		from core.forms import SelectedSeatEditForm
		user = LandingFormData.objects.get(id=user_id)
		if request.method == 'POST':
			form = SelectedSeatEditForm(request.POST, request.FILES, user=user)
			if form.is_valid():
				# Update selected seats
				SelectedSeat.objects.filter(user=user).delete()
				for seat in form.cleaned_data['seats']:
					SelectedSeat.objects.create(user=user, seat=seat, price=seat.price)
				# Optionally update total_paid (if you want to store it elsewhere)
				from django.urls import reverse
				return redirect(reverse('admin:selectedseat-summary'))
		else:
			form = SelectedSeatEditForm(user=user)
		template = loader.get_template("admin/edit_selectedseat.html")
		context = {"form": form, "back_url": "/admin/core/landingformdata/selectedseat-summary/"}
		return HttpResponse(template.render(context, request))

	def selectedseat_summary_view(self, request):
		users = LandingFormData.objects.all()
		report = []
		for user in users:
			seats = SelectedSeat.objects.filter(user=user)
			seat_labels = [s.seat.seat_number for s in seats]
			total_paid = sum(float(s.price or 0) for s in seats)
			earliest = seats.order_by('selected_at').first().selected_at if seats.exists() else None
			# Get latest payment screenshot for user (if any)
			screenshot_url = None
			payment = user.payments.order_by('-uploaded_at').first() if hasattr(user, 'payments') else None
			if payment and payment.image:
				screenshot_url = payment.image.url
			report.append({
				'id': user.id,
				'user': user.name,
				'phone': user.phone,
				'seats': ', '.join(seat_labels),
				'total_paid': total_paid,
				'selected_at': earliest,
				'screenshot_url': screenshot_url,
			})
		template = loader.get_template("admin/selectedseat_summary.html")
		context = {"report": report}
		return HttpResponse(template.render(context, request))

admin.site.register_view = lambda name, view, url=None: admin.site.get_urls().insert(0, path(url or name + '/', view, name=name))
from django.contrib import admin
from django.urls import path
from django.http import HttpResponse
from django.template import loader
from core.models import SelectedSeat, LandingFormData

class BookingReportAdmin(admin.ModelAdmin):
	change_list_template = "admin/booking_report.html"

	def get_urls(self):
		urls = super().get_urls()
		custom_urls = [
			path('booking-report/', self.admin_site.admin_view(self.booking_report_view), name="booking-report"),
		]
		return custom_urls + urls

	def booking_report_view(self, request):
		# Aggregate booking info per user
			users = LandingFormData.objects.all()
			report = []
			for user in users:
				seats = SelectedSeat.objects.filter(user=user)
				seat_labels = [s.seat.seat_number for s in seats]
				total_paid = sum(float(s.price or 0) for s in seats)
				earliest = seats.order_by('selected_at').first().selected_at if seats.exists() else None
				report.append({
					'id': user.id,
					'user': user.name,
					'phone': user.phone,
					'seats': ', '.join(seat_labels),
					'total_paid': total_paid,
					'selected_at': earliest,
				})
			template = loader.get_template("admin/booking_report.html")
			context = {"report": report}
			return HttpResponse(template.render(context, request))

## Removed lambda-based admin.site.register_view code

from django.contrib import admin
from .models import LandingFormData, PaymentScreenshot, Seat, SelectedSeat


@admin.register(Seat)
class SeatAdmin(admin.ModelAdmin):
	list_display = ('seat_number', 'status', 'is_booked', 'reserved_by_name')
	list_filter = ('status', 'is_booked')
	search_fields = ('seat_number',)
	actions = ['mark_as_booked', 'mark_as_unbooked']

	def reserved_by_name(self, obj):
		return obj.reserved_by.name if obj.reserved_by else "-"
	reserved_by_name.short_description = 'Reserved By'

	def mark_as_booked(self, request, queryset):
		# Update both status and is_booked flag
		updated = queryset.update(status='booked', is_booked=True)
		self.message_user(request, f"{updated} seat(s) marked as booked.")
	mark_as_booked.short_description = "Mark selected seats as booked"

	def mark_as_unbooked(self, request, queryset):
		# Update both status and is_booked flag, and clear reservation info
		updated = queryset.update(
			status='available', 
			is_booked=False, 
			reserved_by=None, 
			reserved_until=None
		)
		self.message_user(request, f"{updated} seat(s) marked as available.")
	mark_as_unbooked.short_description = "Mark selected seats as available (unbooked)"

## Removed duplicate registration for LandingFormData; now only registered with SelectedSeatSummaryAdmin




## Removed default SelectedSeatAdmin registration to avoid duplicate user rows
admin.site.register(LandingFormData, SelectedSeatSummaryAdmin)
